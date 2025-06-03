import pandas as pd
import io
import sys
import json
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def toJSON(results, options):
    """Converts the results to JSON format and writes it to a file or prints it to the console."""

    if 'filename' in options and options['filename'] is not None:
        with open(options['filename'], 'w') as file:
            json.dump(results, file)
    else:
        print(json.dumps(results, indent=2))

def toExcel(content, options):
    """Converts the results to an Excel file and writes it to a file or prints it to the console."""
    df = pd.DataFrame(content)
    citations = pd.DataFrame(df['citations'].tolist())
    df = df.drop('citations', axis=1)
    df = pd.concat([df[['question', 'answer']], citations], axis=1)
    df.columns = ['Question', 'Answer'] + [f'Cite {i+1}' for i in range(citations.shape[1])]

    if 'filename' in options and options['filename'] is not None:
        filename = options['filename']
        df.to_excel(filename)
        resizeExcel(filename)
        print("Data written to file: ", filename)
    else:
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer) as writer:
            df.to_excel(writer)
        buffer.seek(0)
        sys.stdout.buffer.write(buffer.read())

def toRaw(content, _options):
    """Prints the raw content to the console."""
    print(content)

def resizeExcel(filename):
    """Resizes the columns and rows in the Excel file."""
    workbook = load_workbook(filename)
    sheet = workbook.active

    sheet.column_dimensions['B'].width = 140
    sheet.column_dimensions['C'].width = 30

    for row in sheet.iter_rows():
        for cell in row:
            if cell.column_letter in ['B', 'C']:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                sheet.row_dimensions[cell.row].height = None
            else:
                cell.alignment = Alignment(vertical='top')

    workbook.save(filename)
